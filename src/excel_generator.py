# ==============================================================
# EXCEL GENERATOR
# ==============================================================
# Gera automaticamente a Matriz de Planejamento da Auditoria
# ==============================================================

import os
import re
import json
from openpyxl import load_workbook

from src.rag_engine import (
    get_or_create_vector_db,
    search_norms
)
from src.llm_cache import buscar_cache, salvar_cache

from src.config import (
    GENERAL_NORMS_PATH,
    VECTOR_DB_NAME
)

# ==============================================================
# CONFIGURAÇÃO DE TAMANHO DE LOTE
# ==============================================================
BATCH_SIZE = 1

# ==============================================================
# LIMPAR NOME DO ARQUIVO
# ==============================================================
def limpar_nome_arquivo(texto):
    texto = texto.strip().replace(" ", "_")
    texto = re.sub(r'[\\/*?:"<>|]', "", texto)
    return texto

# ==============================================================
# SUBSTITUIR PLACEHOLDERS DO TEMPLATE
# ==============================================================
def substituir_placeholders(ws, dados):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for chave, valor in dados.items():
                    placeholder = "{{" + chave + "}}"
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(
                            placeholder,
                            str(valor)
                        )

# ==============================================================
# GERAR MATRIZ DE PLANEJAMENTO
# ==============================================================
def gerar_matriz_planejamento(
        resultado_llm,
        dados_auditoria,
        riscos,
        project_dir,
        chamar_llm):

    from src.audit_generator import extrair_json
    
    template_path = "templates/matriz_planejamento_template.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    # ----------------------------------------------------------
    # AJUSTE: SUBSTITUIR PLACEHOLDERS COM DADOS DA INTERFACE
    # ----------------------------------------------------------
    dados_template = {
        "objetivo": dados_auditoria.get("objetivo", ""),
        "unidades_auditadas": dados_auditoria.get("unidades_auditadas", ""),
        "equipe": dados_auditoria.get("equipe") or "Equipe de Auditoria"
    }

    substituir_placeholders(ws, dados_template)

    # ----------------------------------------------------------
    # PREPARAR RAG
    # ----------------------------------------------------------
    project_norms_path = os.path.join(project_dir, "normas")
    project_vector_path = os.path.join(project_dir, VECTOR_DB_NAME)
    general_vector_path = os.path.join("data", VECTOR_DB_NAME)

    print("\nInicializando RAG...")

    project_db = get_or_create_vector_db(
        project_norms_path,
        project_vector_path
    )

    general_db = get_or_create_vector_db(
        GENERAL_NORMS_PATH,
        general_vector_path
    )

    # ----------------------------------------------------------
    # DADOS GERADOS PELO LLM
    # ----------------------------------------------------------
    questoes = resultado_llm.get("questoes_auditoria", [])
    procedimentos = resultado_llm.get("procedimentos_auditoria", [])

    # ----------------------------------------------------------
    # COLETAR TODOS OS PROCEDIMENTOS
    # ----------------------------------------------------------
    procedimentos_flat = []
    for i, bloco in enumerate(procedimentos):
        questao = questoes[i] if i < len(questoes) else ""
        risco = ""
        if len(riscos) > 0:
            risco = riscos[i % len(riscos)].get("risco", "")

        numero_questao = i + 1
        for j, proc in enumerate(bloco.get("procedimentos", [])):
            numero_subquestao = f"{numero_questao}.{j+1}"
            contexto_normas = search_norms(
                proc,
                project_db,
                general_db
            )
            procedimentos_flat.append({
                "risco": risco,
                "questao": questao,
                "numero": numero_subquestao,
                "procedimento": proc,
                "contexto_normas": contexto_normas
            })

    # ----------------------------------------------------------
    # PROMPT BASE (IMPORTANTE!!)
    # ----------------------------------------------------------
    prompt_base = """
Você é auditor governamental.

Preencha os campos de uma MATRIZ DE PLANEJAMENTO DE AUDITORIA usando:

- risco identificado
- procedimento de auditoria
- trechos de normas recuperados pelo RAG

Regras:

1. Gere exatamente um resultado para cada procedimento.
2. Mantenha a mesma ordem recebida.
3. Responda SOMENTE com JSON válido.
4. Não escreva texto antes ou depois do JSON.
- Importante: Se precisar usar aspas dentro de um texto, use aspas simples (') ou escape-as (\").

Campos:

CRITÉRIO
Deve citar norma + número + artigo/inciso + obrigação normativa.
Exemplo:
Lei 12.527/2011, art. 8º – divulgação ativa de informações.
IN TCU 84/2020, arts. 8º e 9º – divulgação de informações de gestão.

INFORMAÇÕES REQUERIDAS
Evidências necessárias (documentos, registros de sistema, relatórios, evidências de publicação).

FONTES DE INFORMAÇÃO
Origem das evidências (sistemas institucionais, portal da transparência, documentos administrativos, unidades responsáveis, entrevistas).

POSSÍVEIS LIMITAÇÕES
Fatores que dificultam a auditoria (ausência de registros, acesso restrito a sistemas, documentação incompleta).

POSSÍVEIS ACHADOS
Irregularidades possíveis (ausência de divulgação obrigatória, atraso na publicação, inconsistências, falta de controles).

Formato obrigatório:

{
"resultados":[
{
"criterio":"",
"informacoes_requeridas":"",
"fontes_informacao":"",
"possiveis_limitacoes":"",
"possiveis_achados":""
}
]
}

PROCEDIMENTOS DE AUDITORIA:
"""

    # ----------------------------------------------------------
    # EXECUTAR LLM EM LOTES
    # ----------------------------------------------------------
    resultados = []
    total = len(procedimentos_flat)
    print(f"\nTotal de procedimentos: {total}")

    for inicio in range(0, total, BATCH_SIZE):
        lote = procedimentos_flat[inicio:inicio+BATCH_SIZE]
        prompt = prompt_base
        for item in lote:
            prompt += f"""
RISCO IDENTIFICADO:
{item["risco"]}

PROCEDIMENTO DE AUDITORIA:
{item["procedimento"]}

TRECHOS DE NORMAS RECUPERADOS PELO RAG:
{item["contexto_normas"]}
--------------------------------
"""
        cache = buscar_cache(prompt)
        if cache:
            print("Resposta da matriz recuperada do cache.")
            resposta = cache
        else:
            resposta = chamar_llm(prompt)
            salvar_cache(prompt, resposta)

        try:
            # Usamos a função robusta que já limpa caracteres invisíveis e aspas mal formatadas
            dados = extrair_json(resposta)
            
            # Garante que resultados_lote seja uma lista, mesmo em caso de erro na extração
            resultados_lote = dados.get("resultados", [])
            
            if not resultados_lote:
                 print(f"\n[Aviso] Lote {inicio} retornou JSON válido, mas sem a chave 'resultados'.")

            resultados.extend(resultados_lote)

        except Exception as e:
            print("\nErro ao interpretar resposta do LLM no Excel Generator.")
            print(f"Erro técnico: {e}")
            # Em vez de travar toda a aplicação, adiciona campos vazios para este lote
            for _ in lote:
                resultados.append({
                    "criterio": "Erro na geração (Verificar logs)",
                    "informacoes_requeridas": "",
                    "fontes_informacao": "",
                    "possiveis_limitacoes": "",
                    "possiveis_achados": ""
                })

    # ----------------------------------------------------------
    # PREENCHER MATRIZ E SALVAR
    # ----------------------------------------------------------
    linha = 9
    for i, item in enumerate(procedimentos_flat):
        campos = resultados[i] if i < len(resultados) else {}
        ws.cell(row=linha, column=1).value = item["risco"]
        ws.cell(row=linha, column=2).value = item["questao"]
        ws.cell(row=linha, column=3).value = item["numero"]
        ws.cell(row=linha, column=4).value = item["procedimento"]
        ws.cell(row=linha, column=5).value = campos.get("criterio", "")
        ws.cell(row=linha, column=6).value = campos.get("informacoes_requeridas", "")
        ws.cell(row=linha, column=7).value = campos.get("fontes_informacao", "")
        ws.cell(row=linha, column=8).value = campos.get("possiveis_limitacoes", "")
        ws.cell(row=linha, column=11).value = campos.get("possiveis_achados", "")
        linha += 1

    titulo = dados_auditoria["titulo_auditoria"]
    nome_limpo = limpar_nome_arquivo(titulo)
    nome_arquivo = f"Matriz_Planejamento_{nome_limpo}.xlsx"
    output_dir = os.path.join(project_dir, "output")
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    caminho_saida = os.path.join(output_dir, nome_arquivo)
    wb.save(caminho_saida)

    return caminho_saida